"""
API views для управления производственными данными
"""
from rest_framework import generics, status, filters
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.http import HttpResponse
from django.core.management import call_command
import logging
import csv
import json
from datetime import datetime
from io import StringIO

from .models import ProductionLine, PlanTask, Downtime
from .serializers import (
    ProductionLineSerializer, 
    PlanTaskSerializer, 
    DowntimeSerializer,
    PlanUploadSerializer
)

logger = logging.getLogger(__name__)


class ProductionLineListView(generics.ListCreateAPIView):
    """Список производственных линий"""
    
    queryset = ProductionLine.objects.prefetch_related('aliases').filter(is_active=True)
    serializer_class = ProductionLineSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'aliases__alias']
    ordering = ['name']


class ProductionLineDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Детали производственной линии"""
    
    queryset = ProductionLine.objects.prefetch_related('aliases')
    serializer_class = ProductionLineSerializer


class PlanTaskListView(generics.ListCreateAPIView):
    """Список задач плана производства"""
    
    queryset = PlanTask.objects.select_related('production_line', 'product')
    serializer_class = PlanTaskSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['production_line', 'source']
    ordering = ['start_dt', 'production_line__name']
    
    def get_queryset(self):
        """Фильтрация по датам"""
        queryset = super().get_queryset()
        
        # Фильтрация по диапазону дат
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%d-%m-%Y').date()
                queryset = queryset.filter(start_dt__gte=start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                from datetime import datetime
                end_dt = datetime.strptime(end_date, '%d-%m-%Y').date()
                queryset = queryset.filter(end_dt__lte=end_dt)
            except ValueError:
                pass
        
        return queryset


class PlanTaskDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Детали задачи плана"""
    
    queryset = PlanTask.objects.select_related('production_line', 'product')
    serializer_class = PlanTaskSerializer


class DowntimeListView(generics.ListCreateAPIView):
    """Список простоев производственных линий"""
    
    queryset = Downtime.objects.select_related('line')
    serializer_class = DowntimeSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['line', 'status', 'kind', 'source']
    ordering = ['-confidence', 'start_dt']
    
    def get_queryset(self):
        """Фильтрация по датам и уверенности"""
        queryset = super().get_queryset()
        
        # Фильтрация по диапазону дат
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            try:
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%d-%m-%Y').date()
                queryset = queryset.filter(start_dt__gte=start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                from datetime import datetime
                end_dt = datetime.strptime(end_date, '%d-%m-%Y').date()
                queryset = queryset.filter(end_dt__lte=end_dt)
            except ValueError:
                pass
        
        # Фильтрация по минимальной уверенности
        min_confidence = self.request.query_params.get('min_confidence')
        if min_confidence:
            try:
                min_conf = float(min_confidence)
                queryset = queryset.filter(confidence__gte=min_conf)
            except ValueError:
                pass
        
        return queryset


class DowntimeDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Детали простоя"""
    
    queryset = Downtime.objects.select_related('line')
    serializer_class = DowntimeSerializer


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def upload_plan(request):
    """
    Загрузка плана производства из Excel файла
    
    Принимает Excel файл с точными русскими заголовками:
    - Произ. Задание
    - Продукт
    - Начало выполнения (DD.MM.YYYY)
    - Завершение выполнения (DD.MM.YYYY)
    """
    try:
        # Проверка размера файла (20 МБ максимум)
        if hasattr(request, 'FILES') and 'file' in request.FILES:
            uploaded_file = request.FILES['file']
            max_size = 20 * 1024 * 1024  # 20 МБ
            if uploaded_file.size > max_size:
                return Response({
                    'success': False,
                    'data': None,
                    'error': {
                        'code': 'PAYLOAD_TOO_LARGE',
                        'message': f'Размер файла ({uploaded_file.size / 1024 / 1024:.1f} МБ) превышает максимально допустимый (20 МБ)',
                        'details': {
                            'file_size_mb': round(uploaded_file.size / 1024 / 1024, 1),
                            'max_size_mb': 20
                        }
                    }
                }, status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE)
        
        # Валидация с помощью сериализатора
        serializer = PlanUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'data': None,
                'error': {
                    'code': 'VALIDATION_ERROR',
                    'message': 'Ошибка валидации загруженного файла',
                    'details': serializer.errors
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = serializer.validated_data['file']
        
        # Логируем загрузку
        logger.info(f"Plan upload started: {uploaded_file.name}, size: {uploaded_file.size}")
        
        # Интеграция с обработчиком Excel файлов
        from apps.documents.services import FileProcessingManager
        import asyncio
        
        processor = FileProcessingManager()
        
        # Запускаем асинхронную обработку
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            processing_result = loop.run_until_complete(processor.process_file(uploaded_file))
        finally:
            loop.close()
        
        if not processing_result.success:
            error_message = (
                processing_result.errors[0] if processing_result.errors 
                else "Неизвестная ошибка при обработке файла"
            )
            
            return Response({
                'success': False,
                'data': None,
                'error': {
                    'code': 'FILE_PROCESSING_ERROR',
                    'message': error_message,
                    'details': {
                        'filename': uploaded_file.name,
                        'errors': processing_result.errors,
                        'processing_time_ms': processing_result.processing_time_ms
                    }
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Формируем ответ с актуальными данными
        result_data = {
            'created': processing_result.metadata.get('tasks_created', 0),
            'updated': processing_result.metadata.get('tasks_updated', 0), 
            'warnings': processing_result.warnings,
            'processing_time_ms': processing_result.processing_time_ms,
            'file_hash': processing_result.file_hash
        }
        
        logger.info(
            f"Plan upload completed: {result_data['created']} created, "
            f"{result_data['updated']} updated, {len(result_data['warnings'])} warnings"
        )
        
        return Response({
            'success': True,
            'data': result_data,
            'error': None
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Plan upload failed: {str(e)}", exc_info=True)
        
        return Response({
            'success': False,
            'data': None,
            'error': {
                'code': 'GENERAL_ERROR',
                'message': 'Произошла внутренняя ошибка при загрузке плана',
                'details': {'error_message': str(e)}
            }
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_conflicts(request):
    """
    Получение списка конфликтов между задачами плана и простоями
    """
    try:
        # Получаем все задачи и простои
        tasks = PlanTask.objects.select_related('production_line', 'product').all()
        downtimes = Downtime.objects.select_related('line').all()
        
        conflicts = []
        
        # Поиск пересечений
        for task in tasks:
            for downtime in downtimes:
                # Проверяем только если линии совпадают
                if downtime.line_id == task.production_line_id:
                    # Проверяем пересечение дат (включительно)
                    if (task.start_dt <= downtime.end_dt and 
                        task.end_dt >= downtime.start_dt):
                        
                        # Вычисляем пересечение
                        overlap_start = max(task.start_dt, downtime.start_dt)
                        overlap_end = min(task.end_dt, downtime.end_dt)
                        
                        conflict = {
                            'id': f"conflict_{task.id}_{downtime.id}",
                            'level': 'warning',
                            'code': 'CONFLICT_DETECTED',
                            'text': (
                                f"Конфликт расписания: задача '{task.title}' "
                                f"пересекается с простоем {downtime.line.name if downtime.line else 'Unknown'} "
                                f"с {downtime.start_dt.strftime('%d-%m-%Y')} по {downtime.end_dt.strftime('%d-%m-%Y')}"
                            ),
                            'plan_task': PlanTaskSerializer(task).data,
                            'downtime': DowntimeSerializer(downtime).data,
                            'overlap_start': overlap_start.strftime('%d-%m-%Y'),
                            'overlap_end': overlap_end.strftime('%d-%m-%Y'),
                            'priority_status': downtime.status or 'unknown',
                            'created_at': timezone.now().isoformat()
                        }
                        conflicts.append(conflict)
        
        logger.info(f"Conflicts check completed: {len(conflicts)} conflicts found")
        
        return Response({
            'success': True,
            'data': conflicts,
            'error': None
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Conflicts check failed: {str(e)}", exc_info=True)
        
        return Response({
            'success': False,
            'data': None,
            'error': {
                'code': 'GENERAL_ERROR',
                'message': 'Не удалось получить список конфликтов',
                'details': {'error': str(e)}
            }
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _get_conflicts_data():
    """
    Получение данных о конфликтах для экспорта
    """
    # Получаем все задачи и простои
    tasks = PlanTask.objects.select_related('production_line', 'product').all()
    downtimes = Downtime.objects.select_related('line').all()
    
    conflicts = []
    
    # Поиск пересечений
    for task in tasks:
        for downtime in downtimes:
            # Проверяем только если линии совпадают
            if downtime.line_id == task.production_line_id:
                # Проверяем пересечение дат (включительно)
                if (task.start_dt <= downtime.end_dt and 
                    task.end_dt >= downtime.start_dt):
                    
                    # Вычисляем пересечение
                    overlap_start = max(task.start_dt, downtime.start_dt)
                    overlap_end = min(task.end_dt, downtime.end_dt)
                    
                    conflict = {
                        'id': f"conflict_{task.id}_{downtime.id}",
                        'линия': downtime.line.name if downtime.line else 'Неизвестная линия',
                        'задача_плана': task.title,
                        'продукт': task.product.name if task.product else '',
                        'план_начало': task.start_dt.strftime('%d-%m-%Y'),
                        'план_окончание': task.end_dt.strftime('%d-%m-%Y'),
                        'простой_начало': downtime.start_dt.strftime('%d-%m-%Y'),
                        'простой_окончание': downtime.end_dt.strftime('%d-%m-%Y'),
                        'пересечение_начало': overlap_start.strftime('%d-%m-%Y'),
                        'пересечение_окончание': overlap_end.strftime('%d-%m-%Y'),
                        'тип_простоя': downtime.kind or '',
                        'статус': downtime.status or '',
                        'источник': downtime.source or '',
                        'уверенность': f"{downtime.confidence:.2f}" if downtime.confidence is not None else '',
                        'цитата': downtime.evidence_quote or '',
                        'файл_источника': downtime.source_file or '',
                        'создано': timezone.now().strftime('%d-%m-%Y %H:%M:%S')
                    }
                    conflicts.append(conflict)
    
    return conflicts


@api_view(['GET'])
def export_conflicts_csv(request):
    """
    Экспорт конфликтов в CSV формате
    """
    try:
        conflicts = _get_conflicts_data()
        
        # Создаем CSV
        output = StringIO()
        
        if conflicts:
            fieldnames = list(conflicts[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(conflicts)
        else:
            # Пустой CSV с заголовками
            fieldnames = [
                'id', 'линия', 'задача_плана', 'продукт', 'план_начало', 'план_окончание',
                'простой_начало', 'простой_окончание', 'пересечение_начало', 'пересечение_окончание',
                'тип_простоя', 'статус', 'источник', 'уверенность', 'цитата', 'файл_источника', 'создано'
            ]
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8'
        )
        
        filename = f"конфликты_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Conflicts exported to CSV: {len(conflicts)} conflicts")
        
        return response
        
    except Exception as e:
        logger.error(f"CSV export failed: {str(e)}", exc_info=True)
        return HttpResponse(
            f"Ошибка при экспорте: {str(e)}",
            status=500,
            content_type='text/plain; charset=utf-8'
        )


@api_view(['GET'])
def export_conflicts_json(request):
    """
    Экспорт конфликтов в JSON формате
    """
    try:
        conflicts = _get_conflicts_data()
        
        export_data = {
            'экспортировано': timezone.now().strftime('%d-%m-%Y %H:%M:%S'),
            'всего_конфликтов': len(conflicts),
            'конфликты': conflicts
        }
        
        response = HttpResponse(
            json.dumps(export_data, ensure_ascii=False, indent=2),
            content_type='application/json; charset=utf-8'
        )
        
        filename = f"конфликты_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Conflicts exported to JSON: {len(conflicts)} conflicts")
        
        return response
        
    except Exception as e:
        logger.error(f"JSON export failed: {str(e)}", exc_info=True)
        return HttpResponse(
            json.dumps({
                'ошибка': f"Не удалось экспортировать данные: {str(e)}"
            }, ensure_ascii=False),
            status=500,
            content_type='application/json; charset=utf-8'
        )


@api_view(['GET'])
def export_plan_excel(request):
    """
    Экспорт плана производства в Excel формате
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # Получаем данные плана
        queryset = PlanTask.objects.select_related('production_line', 'product')
        
        # Применяем фильтры из query params
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        line_id = request.query_params.get('line_id')
        
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%d-%m-%Y').date()
                queryset = queryset.filter(start_dt__gte=start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%d-%m-%Y').date()
                queryset = queryset.filter(end_dt__lte=end_dt)
            except ValueError:
                pass
        
        if line_id:
            try:
                queryset = queryset.filter(production_line_id=int(line_id))
            except ValueError:
                pass
        
        plan_tasks = queryset.order_by('start_dt', 'production_line__name')
        
        # Создаем Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "План производства"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Заголовки
        headers = [
            'ID', 'Производственная линия', 'Продукт', 'Код продукта',
            'Задача', 'Дата начала', 'Дата окончания', 'Длительность (дни)',
            'Источник', 'Дата создания'
        ]
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Записываем данные
        for row_idx, task in enumerate(plan_tasks, 2):
            data = [
                task.id,
                task.production_line.name if task.production_line else '',
                task.product.name if task.product else '',
                task.product.code if task.product else '',
                task.title,
                task.start_dt.strftime('%d-%m-%Y'),
                task.end_dt.strftime('%d-%m-%Y'),
                task.duration_days,
                task.get_source_display(),
                task.created_at.strftime('%d-%m-%Y %H:%M:%S')
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                # Выравнивание для дат и чисел
                if col_idx in [1, 6, 7, 8, 10]:  # ID, даты, длительность, дата создания
                    cell.alignment = Alignment(horizontal="center")
        
        # Автоподбор ширины колонок
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in worksheet[column_letter]:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            
            # Устанавливаем ширину с небольшим запасом
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем в память
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        # Создаем HTTP ответ
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"план_производства_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Plan exported to Excel: {len(plan_tasks)} tasks")
        
        return response
        
    except ImportError:
        logger.error("openpyxl library is not installed")
        return HttpResponse(
            "Библиотека openpyxl не установлена. Невозможно экспортировать в Excel.",
            status=500,
            content_type='text/plain; charset=utf-8'
        )
        
    except Exception as e:
        logger.error(f"Excel export failed: {str(e)}", exc_info=True)
        return HttpResponse(
            f"Ошибка при экспорте в Excel: {str(e)}",
            status=500,
            content_type='text/plain; charset=utf-8'
        )


@api_view(['GET'])  
def export_plan_csv(request):
    """
    Экспорт плана производства в CSV формате
    """
    try:
        # Получаем данные плана с фильтрами
        queryset = PlanTask.objects.select_related('production_line', 'product')
        
        # Применяем фильтры из query params
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        line_id = request.query_params.get('line_id')
        
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%d-%m-%Y').date()
                queryset = queryset.filter(start_dt__gte=start_dt)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%d-%m-%Y').date()
                queryset = queryset.filter(end_dt__lte=end_dt)
            except ValueError:
                pass
        
        if line_id:
            try:
                queryset = queryset.filter(production_line_id=int(line_id))
            except ValueError:
                pass
        
        plan_tasks = queryset.order_by('start_dt', 'production_line__name')
        
        # Создаем CSV
        output = StringIO()
        
        fieldnames = [
            'ID', 'производственная_линия', 'продукт', 'код_продукта',
            'задача', 'дата_начала', 'дата_окончания', 'длительность_дни',
            'источник', 'дата_создания'
        ]
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for task in plan_tasks:
            writer.writerow({
                'ID': task.id,
                'производственная_линия': task.production_line.name if task.production_line else '',
                'продукт': task.product.name if task.product else '',
                'код_продукта': task.product.code if task.product else '',
                'задача': task.title,
                'дата_начала': task.start_dt.strftime('%d-%m-%Y'),
                'дата_окончания': task.end_dt.strftime('%d-%m-%Y'),
                'длительность_дни': task.duration_days,
                'источник': task.get_source_display(),
                'дата_создания': task.created_at.strftime('%d-%m-%Y %H:%M:%S')
            })
        
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8'
        )
        
        filename = f"план_производства_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Plan exported to CSV: {len(plan_tasks)} tasks")
        
        return response
        
    except Exception as e:
        logger.error(f"CSV export failed: {str(e)}", exc_info=True)
        return HttpResponse(
            f"Ошибка при экспорте в CSV: {str(e)}",
            status=500,
            content_type='text/plain; charset=utf-8'
        )
    
@api_view(["POST"])
def reset_database(request):
    try:
        # flush сбрасывает все таблицы и автоинкременты
        call_command("flush", "--noinput")
        return Response({"status": "ok", "message": "Database has been cleared."})
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=500)