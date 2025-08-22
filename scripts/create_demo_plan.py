#!/usr/bin/env python3
"""
Создание демонстрационного Excel файла с планом производства
"""

import sys
import os
from pathlib import Path

# Добавляем корневую директорию проекта в Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_demo_plan_excel():
    """Создание demo/plan/plan.xlsx с точным форматом"""
    
    # Создаем новую книгу
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "План производства"
    
    # Стили для заголовков
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Точные заголовки из требований
    headers = [
        'Произ. Задание',
        'Продукт', 
        'Начало выполнения',
        'Завершение выполнения'
    ]
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Данные согласно требованиям (включая невозможную дату 31.04.2026)
    data = [
        ['1', 'Продукт1', '1.01.2026', '31.04.2026'],
        ['2', 'Продукт2', '31.04.2026', '31.07.2026'],
        ['3', 'Продукт3', '31.07.2026', '31.12.2026']
    ]
    
    # Записываем данные
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Выравнивание для первой колонки (номер задания)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
    
    # Автоподбор ширины колонок
    for col_idx in range(1, len(headers) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        max_length = 0
        
        for row in worksheet[column_letter]:
            if row.value:
                max_length = max(max_length, len(str(row.value)))
        
        # Устанавливаем ширину с запасом
        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 12)
    
    # Сохраняем файл
    output_path = project_root / 'demo' / 'plan' / 'plan.xlsx'
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    workbook.save(str(output_path))
    print(f"✅ Создан файл: {output_path}")
    
    return output_path

if __name__ == "__main__":
    try:
        created_file = create_demo_plan_excel()
        print(f"Демонстрационный файл плана создан: {created_file}")
    except Exception as e:
        print(f"❌ Ошибка создания файла: {e}")
        sys.exit(1)